from openpyxl import load_workbook
import os

def load_excel(path):
    wb = load_workbook(path, data_only=True)
    for ws in wb:
        title = ws["A1"].value
        if "LAYTIME CALCULATION" in title:
            print("Valid worksheet", ws.title)
        vehicle = ws["D2"].value
        vehicle_mother = ws["D3"].value
        discharge_port = ws["B5"].value
        NOR_tendered = ws["G5"].value
        NOR_accepted = ws["G7"].value

        commenced_discharging = discharge_port = ws["G9"].value
        completed_discharging = discharge_port = ws["G11"].value
        cargo_quantity = discharge_port = ws["B9"].value
        discharge_rate = discharge_port = ws["B11"].value
        demurrage = discharge_port = ws["B13"].value
        despatch = discharge_port = ws["B14"].value
        laytime_allowed = discharge_port = ws["B16"].value

        print(vehicle, vehicle_mother, discharge_port, NOR_tendered, NOR_accepted)
        print(commenced_discharging, completed_discharging, cargo_quantity, discharge_rate, demurrage, despatch, laytime_allowed)

        for row in ws.iter_rows(min_row=20, max_row=24):
            print("---")
            for cell in row:
                print(cell.value)

for root, dirs, files in os.walk("./sample", topdown=False):
    for name in files:
        if "DUBAIKNIGHT" in name:
            load_excel(os.path.join(root, name))
