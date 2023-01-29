# -*- encoding: utf-8 -*-
"""
License: MIT
Copyright (c) 2019 - present AppSeed.us
"""
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.template import loader
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django import template
from app.models import Document, Laysheet
from openpyxl import load_workbook

@login_required(login_url="/login/")
def index(request):
    documents = Document.objects.all()
    laysheets = Laysheet.objects.all()
    files = documents.order_by('-uploaded_at')[:10]
    laysheetId = request.GET.get('laysheetId', '')
    if laysheetId:
        laysheets = Laysheet.objects.filter(document=int(laysheetId))
    context = {
        "file_count": documents.count(),
        "laysheets_count": laysheets.count(),
        "current_laysheets": laysheets,
        "files": files
    }
    return render(request, "index.html", context)

@login_required(login_url="/login/")
def pages(request):
    context = {}
    try:
        html_template = loader.get_template(request.path.split('/')[-1])
    except template.TemplateDoesNotExist:
        html_template = loader.get_template( 'error-404.html' )
    except:
        html_template = loader.get_template( 'error-500.html' )
    finally:
        return HttpResponse(html_template.render(context, request))

def load_excel(path):
    wb = load_workbook(path, data_only=True)
    for ws in wb:
        if ws["A1"].value and "BẢNG TÍNH THỜI GIAN DÔI NHẬT" in ws["A1"].value:
            print("valid worksheet", ws.title)
        else:
            print("invalid worksheet", ws.title)
            continue
        vehicle = ws["E2"].value
        if vehicle == "#N/A":
            continue
        vehicle_trip = ws["E3"].value
        contract = ws["E4"].value
        discharge_port = ws["B7"].value
        arrival_windows = ws["B9"].value
        NOR_tendered = ws["H7"].value
        NOR_accepted = ws["H9"].value
        commenced_discharging = ws["H11"].value
        completed_discharging = ws["H13"].value
        weight = ws["B11"].value
        load = ws["B13"].value
        despatch = ws["B15"].value
        demurrage = ws["B17"].value
        laytime_allowed = ws["A16"].value

        real_despatch = 0
        real_demurrage  = 0
        for row in ws.iter_rows(min_row=20):
            for cell in row:
                if "Tiền thưởng" in str(cell.value):
                    real_despatch = ws["H" + str(cell.row)].value
                if "Tiền phạt" in str(cell.value):
                    real_demurrage = ws["H" + str(cell.row)].value

        results = {
            "vehicle": vehicle,
            "vehicle_trip": vehicle_trip,
            "contract": contract,
            "discharge_port": discharge_port,
            "NOR_tendered": NOR_tendered,
            "NOR_accepted": NOR_accepted,
            "commenced_discharging": commenced_discharging,
            "completed_discharging": completed_discharging,
            "weight": weight,
            "load": load,
            "despatch": despatch,
            "demurrage": demurrage,
            "laytime_allowed": laytime_allowed,
            "real_despatch": real_despatch,
            "real_demurrage": real_demurrage
        }
        yield results

@login_required(login_url="/login/")
@csrf_exempt
def document_upload(request):
    if request.method == 'POST':
        document = request.FILES.get("file")
        document_title = request.POST.get("title")
        fs = FileSystemStorage()
        filename = fs.save(document_title, document)
        db_document = Document.objects.create(
            title = document_title,
            path = filename
        )
        documents_info = load_excel(fs.path(filename))
        for document_info in documents_info:
            print(document_info)
            # laysheet = Laysheet.objects.create(
            #     document=db_document,
            #     **document_info
            # )

    return HttpResponse("success")