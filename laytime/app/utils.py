from openpyxl import load_workbook
import os

def load_excel(path):
    wb = load_workbook(path, data_only=True)
    for ws in wb:
        if ws["A1"].value and "BẢNG TÍNH THỜI GIAN DÔI NHẬT" in ws["A1"].value:
            print("valid worksheet", ws.title)
        else:
            print("invalid worksheet", ws.title)
            continue
        vehicle = ws["E2"].value
        if vehicle == "#N/A":
            continue
        vehicle_trip = ws["E3"].value
        contract = ws["E4"].value
        discharge_port = ws["B7"].value
        arrival_windows = ws["B9"].value
        NOR_tendered = ws["H7"].value
        NOR_accepted = ws["H9"].value
        commenced_discharging = ws["H11"].value
        completed_discharging = ws["H13"].value
        weight = ws["B11"].value
        load = ws["B13"].value
        despatch = ws["B15"].value
        demurrage = ws["B17"].value
        laytime_allowed = ws["H15"].value

        real_despatch = 0
        real_demurrage  = 0
        for row in ws.iter_rows(min_row=20):
            for cell in row:
                if "Tiền thưởng" in str(cell.value):
                    real_despatch = ws["H" + str(cell.row)].value
                    if real_despatch == None:
                        real_despatch = 0
                    elif real_despatch == "#REF!":
                        real_despatch = -1
                    elif real_despatch == "#NUM!":
                        real_despatch = -1
                if "Tiền phạt" in str(cell.value):
                    real_demurrage = ws["H" + str(cell.row)].value
                    if real_demurrage == None:
                        real_demurrage = 0
                    elif real_demurrage == "#REF!":
                        real_demurrage = -1
                    elif real_demurrage == "#NUM!":
                        real_demurrage = -1

        results = {
            "vehicle": vehicle,
            "vehicle_trip": vehicle_trip,
            "contract": contract,
            "arrival_windows": arrival_windows,
            "discharge_port": discharge_port,
            "NOR_tendered": NOR_tendered,
            "NOR_accepted": NOR_accepted,
            "commenced_discharging": commenced_discharging,
            "completed_discharging": completed_discharging,
            "weight": weight,
            "load": load,
            "despatch": despatch,
            "demurrage": demurrage,
            "laytime_allowed": laytime_allowed,
            "real_despatch": real_despatch,
            "real_demurrage": real_demurrage
        }
        yield results

def delete_document(path):
    if os.path.isfile(path):
        os.remove(path)
